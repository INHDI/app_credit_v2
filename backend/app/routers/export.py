"""
Export API routes - Export data to Excel
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from app.core.database import get_db
from app.core.deps import require_admin
from app.models.user import User
from app.models.tin_chap import TinChap
from app.models.tra_gop import TraGop

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/export",
    tags=["Export"]
)


def format_currency(value: int) -> str:
    """Format number as Vietnamese currency"""
    if value is None:
        return "0"
    return f"{value:,}".replace(",", ".")


def format_date(date_value) -> str:
    """Format date to Vietnamese format"""
    if date_value is None:
        return ""
    if isinstance(date_value, str):
        return date_value
    return date_value.strftime("%d/%m/%Y")


def style_header(ws, row_num: int):
    """Apply header styling to a row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def style_data_cells(ws, start_row: int, end_row: int, num_cols: int):
    """Apply styling to data cells"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")


@router.get("/tin-chap")
async def export_tin_chap(
    status: Optional[str] = None,
    search: Optional[str] = None,
    ma_hd: Optional[str] = None,
    ho_ten_list: Optional[str] = None,  # Comma-separated list of HoTen
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """
    Export Tin Chap contracts to Excel
    
    - **status**: Filter by status (optional)
    - **search**: Search by name or MaHD (optional)
    - **ma_hd**: Export single contract by MaHD (optional)
    - **ho_ten_list**: Export by specific debtor names, comma-separated (optional)
    """
    # Build query
    query = db.query(TinChap)
    
    if ma_hd:
        query = query.filter(TinChap.MaHD == ma_hd)
    else:
        if ho_ten_list:
            # Filter by list of names
            names = [name.strip() for name in ho_ten_list.split(',')]
            query = query.filter(TinChap.HoTen.in_(names))
        if status:
            query = query.filter(TinChap.TrangThai.contains(status))
        if search:
            query = query.filter(
                (TinChap.MaHD.ilike(f"%{search}%")) |
                (TinChap.HoTen.ilike(f"%{search}%"))
            )
    
    contracts = query.order_by(TinChap.MaHD.desc()).all()
    logger.info(f"📊 Export TinChap: Found {len(contracts)} contracts (ho_ten_list: {ho_ten_list})")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TinChap"  # Use ASCII name for better compatibility
    
    # Add title
    ws.merge_cells('A1:H1')
    ws['A1'] = "DANH SÁCH HỢP ĐỒNG TÍN CHẤP"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add export date
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = ["STT", "Mã HĐ", "Họ Tên", "Ngày Vay", "Số Tiền Vay", "Kỳ Đóng", "Lãi Suất", "Trạng Thái"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header(ws, 4)
    
    # Add data
    for idx, contract in enumerate(contracts, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=contract.MaHD)
        ws.cell(row=row, column=3, value=contract.HoTen)
        ws.cell(row=row, column=4, value=format_date(contract.NgayVay))
        ws.cell(row=row, column=5, value=format_currency(contract.SoTienVay) + " đ")
        ws.cell(row=row, column=6, value=f"{contract.KyDong} ngày")
        ws.cell(row=row, column=7, value=format_currency(contract.LaiSuat) + " đ")
        ws.cell(row=row, column=8, value=contract.TrangThai)
    
    if contracts:
        style_data_cells(ws, 5, 4 + len(contracts), len(headers))
    
    # Add summary
    summary_row = 5 + len(contracts) + 1
    ws.cell(row=summary_row, column=1, value=f"Tổng số hợp đồng: {len(contracts)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    total_amount = sum(c.SoTienVay or 0 for c in contracts)
    ws.cell(row=summary_row + 1, column=1, value=f"Tổng số tiền vay: {format_currency(total_amount)} đ")
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    
    # Adjust column widths
    column_widths = [6, 12, 25, 12, 18, 12, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"TinChap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/tra-gop")
async def export_tra_gop(
    status: Optional[str] = None,
    search: Optional[str] = None,
    ma_hd: Optional[str] = None,
    ho_ten_list: Optional[str] = None,  # Comma-separated list of HoTen
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """
    Export Tra Gop contracts to Excel
    
    - **status**: Filter by status (optional)
    - **search**: Search by name or MaHD (optional)
    - **ma_hd**: Export single contract by MaHD (optional)
    - **ho_ten_list**: Export by specific debtor names, comma-separated (optional)
    """
    # Build query
    query = db.query(TraGop)
    
    if ma_hd:
        query = query.filter(TraGop.MaHD == ma_hd)
    else:
        if ho_ten_list:
            # Filter by list of names
            names = [name.strip() for name in ho_ten_list.split(',')]
            query = query.filter(TraGop.HoTen.in_(names))
        if status:
            query = query.filter(TraGop.TrangThai.contains(status))
        if search:
            query = query.filter(
                (TraGop.MaHD.ilike(f"%{search}%")) |
                (TraGop.HoTen.ilike(f"%{search}%"))
            )
    
    contracts = query.order_by(TraGop.MaHD.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TraGop"  # Use ASCII name for better compatibility
    
    # Add title
    ws.merge_cells('A1:I1')
    ws['A1'] = "DANH SÁCH HỢP ĐỒNG TRẢ GÓP"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add export date
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = ["STT", "Mã HĐ", "Họ Tên", "Ngày Vay", "Số Tiền Vay", "Kỳ Đóng", "Số Lần Trả", "Lãi Suất", "Trạng Thái"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header(ws, 4)
    
    # Add data
    for idx, contract in enumerate(contracts, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=contract.MaHD)
        ws.cell(row=row, column=3, value=contract.HoTen)
        ws.cell(row=row, column=4, value=format_date(contract.NgayVay))
        ws.cell(row=row, column=5, value=format_currency(contract.SoTienVay) + " đ")
        ws.cell(row=row, column=6, value=f"{contract.KyDong} ngày")
        ws.cell(row=row, column=7, value=contract.SoLanTra)
        ws.cell(row=row, column=8, value=format_currency(contract.LaiSuat) + " đ")
        ws.cell(row=row, column=9, value=contract.TrangThai)
    
    if contracts:
        style_data_cells(ws, 5, 4 + len(contracts), len(headers))
    
    # Add summary
    summary_row = 5 + len(contracts) + 1
    ws.cell(row=summary_row, column=1, value=f"Tổng số hợp đồng: {len(contracts)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    total_amount = sum(c.SoTienVay or 0 for c in contracts)
    ws.cell(row=summary_row + 1, column=1, value=f"Tổng số tiền vay: {format_currency(total_amount)} đ")
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    
    # Adjust column widths
    column_widths = [6, 12, 25, 12, 18, 12, 12, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"TraGop_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
